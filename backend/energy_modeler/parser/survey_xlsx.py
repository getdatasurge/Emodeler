"""Excel survey-sheet importer (spec Ch 10 / surveyor workflow).

Reads the IWFA-style window survey workbook used in the field — the same sheet
3M dealers exchange with their PMs — and aggregates rows into Face inputs the
engine consumes. Each row in the 'Survey Sheet' tab is one window; the W/H
columns are inches, Compass is one of N/NE/E/SE/S/SW/W/NW (the workbook's
bundled Lists tab pins the allowed values), Glass Color is one of Clear/
Bronze/Grey/Green/Blue, and Building ID lets a portfolio (multi-school)
upload split into one project per building."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

# Width/Height are recorded in inches on the IWFA survey sheet; convert to ft^2.
IN2_PER_FT2 = 144.0

VALID_COMPASS = {"N", "NE", "E", "SE", "S", "SW", "W", "NW", "H"}

# Surveyor 'Glass Color' -> our base_glazings catalog id. Defaults assume a
# dual-pane IGU (the modern commercial baseline). Single-pane stock should be
# corrected per-face in the UI after the import.
GLASS_COLOR_TO_GLAZING_ID: dict[str, str] = {
    "clear": "dbl_clear_3mm_13mmAir",
    "bronze": "dbl_bronze_3mm_13mmAir",
    "grey": "dbl_grey_3mm_13mmAir",
    "gray": "dbl_grey_3mm_13mmAir",
    "green": "dbl_green_3mm_13mmAir",
    "blue": "dbl_blue_3mm_13mmAir",
    # Solar / reflective coatings are usually distinct films, not a base tint;
    # default to clear and let the user adjust.
}

# Header tokens we resolve to column indices. The header row in the live sheet
# uses these exact labels ('Survey Sheet' tab, row 1).
_HEADER_ALIASES = {
    "compass": ("compass",),
    "width": ("w", "width", "width (in)", "w (in)"),
    "height": ("h", "height", "height (in)", "h (in)"),
    "glass_color": ("glass color?", "glass color", "color"),
    "building_id": ("building id", "building", "site"),
    "floor": ("floor #", "floor", "floor number"),
    "map_number": ("map number", "map", "map #"),
    "zone": ("zone",),
    # GC3200 (Solar Gard) handheld SHGC meter — surveyors sample some windows
    # to spot-check the assumed glazing. We aggregate the readings per bucket
    # and flag a face when the measured SHGC diverges materially from the
    # catalog SHGC, so the import isn't silently mis-classifying the glass.
    "gc3200": ("gc3200 reading", "gc3200", "shgc reading"),
}

# Catalog SHGCs the GC3200 cross-check compares against, keyed by base_glazing_id.
# Populated from data/base_glazings.json on first use; kept module-level so the
# cross-check is cheap inside the hot loop.
_CATALOG_SHGC_CACHE: dict[str, float] | None = None


def _catalog_shgc() -> dict[str, float]:
    global _CATALOG_SHGC_CACHE
    if _CATALOG_SHGC_CACHE is None:
        from .. import datastore
        _CATALOG_SHGC_CACHE = {g["id"]: float(g["shgc"]) for g in datastore.base_glazings()}
    return _CATALOG_SHGC_CACHE


# Tolerance for the measured-vs-catalog SHGC cross-check. The GC3200 is rated
# +/-0.03 SHGC, plus glazing-to-glazing variability — 0.06 catches a clearly
# mis-classified glass type without false-flagging normal scatter.
_SHGC_DIVERGENCE_THRESHOLD = 0.06


@dataclass
class SurveyImportRow:
    """One aggregated face — (building × orientation × base glazing) — after
    summing W*H across the rows that match. notes carries the floor / map /
    zone provenance for the audit bundle."""
    orientation: str
    area_sqft: float
    count: int
    base_glazing_id: str
    building_id: str | None = None
    notes: str | None = None


def _norm(s: Any) -> str:
    """Lowercased + stripped — for matching tokens (compass, color, headers)."""
    return str(s).strip().lower() if s is not None else ""


def _strip(s: Any) -> str:
    """Stripped only — preserves case for human-facing values (building names)."""
    return str(s).strip() if s is not None else ""


def _find_columns(header: list[Any]) -> dict[str, int]:
    norm = [_norm(c) for c in header]
    out: dict[str, int] = {}
    for key, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                out[key] = norm.index(alias)
                break
    missing = [k for k in ("compass", "width", "height") if k not in out]
    if missing:
        raise ValueError(
            f"Survey sheet is missing required column(s): {missing}. "
            f"Saw headers: {[h for h in header if h is not None]}"
        )
    return out


def _to_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def _glazing_for_color(color: str) -> str | None:
    """Surveyor color string -> base_glazing_id, or None if not recognized."""
    return GLASS_COLOR_TO_GLAZING_ID.get(_norm(color))


def _fmt_set(values: set[str], max_items: int = 8) -> str:
    """Compact human display of an unordered set of small strings."""
    pretty = sorted(values, key=lambda v: (len(v), v))
    if len(pretty) > max_items:
        pretty = pretty[: max_items - 1] + [f"+{len(values) - (max_items - 1)} more"]
    return ", ".join(pretty)


@dataclass
class _Bucket:
    """Working accumulator per (building × orientation × glazing) key."""
    area_sqft: float = 0.0
    count: int = 0
    floors: set[str] = field(default_factory=set)
    maps: set[str] = field(default_factory=set)
    zones: set[str] = field(default_factory=set)
    # GC3200 spot-check sample (one reading per window that the surveyor metered).
    gc3200_readings: list[float] = field(default_factory=list)


def parse_survey_xlsx(
    content: bytes,
    *,
    sheet_name: str = "Survey Sheet",
    default_base_glazing_id: str = "dbl_clear_3mm_13mmAir",
    units: str = "in",
) -> list[SurveyImportRow]:
    """Parse a 3M/IWFA survey workbook and return one SurveyImportRow per
    (building_id × orientation × base_glazing_id) bucket. The caller decides
    whether to flatten across buildings (single project) or split per building.

    Width/Height are inches by default; pass units='ft' to skip the conversion.
    Blank Compass / Color cells fill down from the most recently populated row
    (matches surveyor convention — write the elevation once, list windows
    beneath it). Glass Color maps via GLASS_COLOR_TO_GLAZING_ID; an empty or
    unknown color falls back to default_base_glazing_id."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import survey sheets") from exc

    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        candidates = [n for n in wb.sheetnames if "survey" in n.lower()]
        if not candidates:
            raise ValueError(
                f"Workbook has no sheet matching 'Survey Sheet' (found: {wb.sheetnames})"
            )
        sheet_name = candidates[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    cols = _find_columns(list(header))
    factor = 1.0 if units == "ft" else 1.0 / IN2_PER_FT2

    last_compass = ""
    last_color = ""
    last_building = ""
    # bucket key: (building_id, orientation, glazing_id)
    buckets: dict[tuple[str, str, str], _Bucket] = {}

    def _cell(raw: tuple, key: str) -> Any:
        idx = cols.get(key)
        return raw[idx] if idx is not None and idx < len(raw) else None

    for raw in rows_iter:
        if raw is None or not any(c not in (None, "") for c in raw):
            continue

        compass = _norm(_cell(raw, "compass")).upper()
        if compass:
            last_compass = compass
        compass = compass or last_compass
        if compass not in VALID_COMPASS:
            continue

        color = _norm(_cell(raw, "glass_color"))
        if color:
            last_color = color
        bg_id = _glazing_for_color(last_color) or default_base_glazing_id

        building = _strip(_cell(raw, "building_id"))
        if building:
            last_building = building
        building_id = last_building or None

        w = _to_float(_cell(raw, "width"))
        h = _to_float(_cell(raw, "height"))
        if not w or not h:
            continue
        area_sqft = w * h * factor

        key = (building_id or "", compass, bg_id)
        b = buckets.setdefault(key, _Bucket())
        b.area_sqft += area_sqft
        b.count += 1
        for slot, col in (("floor", b.floors), ("map_number", b.maps), ("zone", b.zones)):
            v = _cell(raw, slot)
            if v not in (None, ""):
                col.add(str(v).strip())
        gc = _to_float(_cell(raw, "gc3200"))
        if gc is not None and 0.0 <= gc <= 1.0:
            b.gc3200_readings.append(gc)

    # Stable orientation order (matches the engine's PEAK_POA table) within
    # each (building, glazing) — keeps imports diff-friendly.
    order = {o: i for i, o in enumerate(["N", "NE", "E", "SE", "S", "SW", "W", "NW", "H"])}
    out: list[SurveyImportRow] = []
    sorted_keys = sorted(
        buckets.keys(), key=lambda k: (k[0], order.get(k[1], 99), k[2])
    )
    catalog_shgc = _catalog_shgc()
    for key in sorted_keys:
        building_id, orientation, glazing_id = key
        b = buckets[key]
        notes_parts: list[str] = []
        if b.floors:
            notes_parts.append(f"Floors: {_fmt_set(b.floors)}")
        if b.maps:
            notes_parts.append(f"Maps: {_fmt_set(b.maps)}")
        if b.zones:
            notes_parts.append(f"Zones: {_fmt_set(b.zones)}")
        if b.gc3200_readings:
            measured = sum(b.gc3200_readings) / len(b.gc3200_readings)
            cat = catalog_shgc.get(glazing_id)
            notes_parts.append(
                f"GC3200 avg SHGC: {measured:.2f} (n={len(b.gc3200_readings)})"
            )
            if cat is not None and abs(measured - cat) > _SHGC_DIVERGENCE_THRESHOLD:
                notes_parts.append(
                    f"REVIEW: measured SHGC {measured:.2f} vs catalog {cat:.2f} "
                    f"for {glazing_id} — likely mis-classified glass"
                )
        out.append(
            SurveyImportRow(
                orientation=orientation,
                area_sqft=round(b.area_sqft, 2),
                count=int(b.count),
                base_glazing_id=glazing_id,
                building_id=building_id or None,
                notes=" · ".join(notes_parts) or None,
            )
        )
    return out


def group_by_building(
    rows: list[SurveyImportRow],
) -> dict[str, list[SurveyImportRow]]:
    """Group survey rows by building_id. Each building's list has one row per
    (orientation × base_glazing_id) — the per-project face list ready to insert."""
    grouped: dict[str, list[SurveyImportRow]] = defaultdict(list)
    for r in rows:
        grouped[r.building_id or "(unnamed)"].append(r)
    return dict(grouped)


def collapse_to_single_project(
    rows: list[SurveyImportRow],
) -> list[SurveyImportRow]:
    """Drop the building dimension; re-aggregate by (orientation, glazing_id).
    Notes from the per-building rows are merged (deduped, '; '-joined)."""
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for r in rows:
        key = (r.orientation, r.base_glazing_id)
        agg = by_key.setdefault(
            key,
            {"area_sqft": 0.0, "count": 0, "buildings": set(), "notes": set()},
        )
        agg["area_sqft"] += r.area_sqft
        agg["count"] += r.count
        if r.building_id:
            agg["buildings"].add(r.building_id)
        if r.notes:
            agg["notes"].add(r.notes)

    order = {o: i for i, o in enumerate(["N", "NE", "E", "SE", "S", "SW", "W", "NW", "H"])}
    out: list[SurveyImportRow] = []
    for (orientation, glazing_id), agg in sorted(
        by_key.items(), key=lambda kv: (order.get(kv[0][0], 99), kv[0][1])
    ):
        parts: list[str] = []
        if agg["buildings"]:
            parts.append(f"Buildings: {_fmt_set(agg['buildings'])}")
        parts.extend(sorted(agg["notes"]))
        out.append(
            SurveyImportRow(
                orientation=orientation,
                area_sqft=round(agg["area_sqft"], 2),
                count=int(agg["count"]),
                base_glazing_id=glazing_id,
                building_id=None,
                notes=" | ".join(parts) or None,
            )
        )
    return out
